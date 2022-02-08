import pyupbit 
import pandas as pd
import datetime 
import time 
import os
from datetime import datetime
from openpyxl import load_workbook

# get access key and secret key for the API
# future updates: store access_key and secret_key in a separate file for better security
access_key = "hkdvoPjvx3GIO7XaEPOG8gQj8txUjyQUjWQdfXKB"
secret_key = "0nx6nJriz6BSnxlxOCzNxPrcy1LKaqBQS4wUFAQf"

# create accessible upbit token using access_key and secret_key
upbit_token = pyupbit.Upbit(access_key, secret_key)

# calculates rsi value for input ohlc dataframe
# ohlc: open, high, low, close value for certain ticker in specific time frame
# period: number of candle bar to be used in rsi calculation
def rsi(ohlc: pd.DataFrame, period: int = 14): 
    delta = ohlc["close"].diff() 
    ups, downs = delta.copy(), delta.copy()
    ups[ups < 0] = 0 
    downs[downs > 0] = 0 

    AU = ups.ewm(com = period-1, min_periods = period).mean() 
    AD = downs.abs().ewm(com = period-1, min_periods = period).mean() 
    RS = AU/AD 
    return pd.Series(100 - (100/(1 + RS)), name = "RSI") 

def by_value(item):
    return item[1]


# buy certain ticker in designated amount (currently fixed 10000)
def buy(ticker):
    upbit_token.buy_market_order(ticker, 10000)

    # Record in bought_list
    bought_tickers = []
    with open('bought_list.txt') as f:
        bought_tickers = f.read().splitlines()
    if ticker not in bought_tickers:
        bought_tickers.append(ticker)
    with open('bought_list.txt', 'w') as output:
        for ticker in bought_tickers:
            output.write(str(ticker) + '\n') 


# check whether certain ticker is in a buy condition
    # buy if both past_rsi and cur_rsi are below 30 and 
    # cur_rsi is greater than past_rsi
def should_buy(ticker, cur_rsi, past_rsi):
    return cur_rsi < 30 and past_rsi < 30 and cur_rsi > past_rsi * 1.01


# sell certain ticker in designated amount
def sell(ticker, selling_balance):
    upbit_token.sell_market_order(ticker, selling_balance)

    # remove ticker from bought_list.txt
    bought_tickers = []
    with open('bought_list.txt') as f:
        bought_tickers = f.read().splitlines()
    bought_tickers.remove(ticker)
    with open('bought_list.txt', 'w') as output:
        for ticker in bought_tickers:
            output.write(str(ticker) + '\n')

# sell if previous rsi value is greater 70 and 
# current rsi is smaller than previous rsi
def should_sell(ticker, cur_rsi, past_rsi):
    if past_rsi > 70 and cur_rsi < past_rsi*0.99:
        bought_tickers = []
        with open('bought_list.txt') as f:
            bought_tickers = f.read().splitlines()

        if ticker in bought_tickers:
            return True
    
    return False

# getting korean tickers
tickers = pyupbit.get_tickers()
kr_tickers = []
past_rsi = {}

for ticker in tickers:
    if ticker[0:3] == "KRW":
        kr_tickers.append(ticker)

for ticker in kr_tickers:
    past_rsi[ticker] = 50

while True:
    for ticker in kr_tickers:
        data = pyupbit.get_ohlcv(ticker = ticker, interval="minute1")
        cur_rsi = rsi(data, 14).iloc[-1]
        
        if should_buy(ticker, cur_rsi, past_rsi[ticker]):
            try:
                buy(ticker)

                current_time = datetime.now().strftime("%H:%M:%S")
                cur_price_per_ticker = pyupbit.get_current_price(ticker)

                new_data_row = [current_time, ticker, 'buy', cur_rsi, past_rsi, '10000', cur_price_per_ticker]

                # Record the buy transaction
                wb = load_workbook('Records.xlsx')
                ws = wb.worksheet[0]
                ws.append(new_data_row)
                wb.save('Records.xlsx')

            except:
                print("error in purchasing " + ticker)

        elif should_sell(ticker, cur_rsi, past_rsi[ticker]):
            try:
                current_time = datetime.now().strftime("%H:%M:%S")
                cur_price_per_ticker = pyupbit.get_current_price(ticker)

                ticker_balance = upbit_token.get_balance(ticker)
                ticker_cur_price = pyupbit.get_current_price(ticker)

                selling_amount_KRW = ticker_balance * ticker_cur_price
                selling_balance = ticker_balance

                if cur_rsi < 75 and selling_amount_KRW > 10000:
                    selling_balance / 2
                
                sell(ticker, selling_balance)

                new_data_row = [current_time, ticker, 'sell', cur_rsi, past_rsi, selling_balance*ticker_cur_price, cur_price_per_ticker]

                # Record the buy transaction
                wb = load_workbook('Records.xlsx')
                ws = wb.worksheet[0]
                ws.append(new_data_row)
                wb.save('Records.xlsx')
            
            except:
                print("error in selling " + ticker)
        
        else:
            print("Current State " + str(ticker) + " at cur_rsi: " + str(cur_rsi) + " and past_rsi: " + str(past_rsi[coin_name]))
        
        past_rsi[ticker] = cur_rsi